"""
Rotas de geração de relatórios (PDF e XLSX)
"""
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
from weasyprint import HTML
from app import db
from app.models.usuario import Usuario
from app.models.atividade import Atividade
from app.models.entrega import Entrega
from app.models.questao import Resposta
from app.utils.auth import professor_required

bp = Blueprint('relatorios', __name__, url_prefix='/api/relatorios')

@bp.route('/desempenho', methods=['GET'])
@professor_required
def relatorio_desempenho():
    """
    Gera relatório de desempenho com:
    - Taxa de entrega
    - Notas médias
    - Ranking
    - Histórico de entregas
    - Estatísticas de múltipla escolha
    """
    turma = request.args.get('turma')
    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')
    formato = request.args.get('formato', 'json')  # json, pdf, xlsx
    
    # Construir query base
    query_alunos = Usuario.query.filter_by(tipo='aluno')
    
    if turma:
        query_alunos = query_alunos.filter_by(turma=turma)
    
    alunos = query_alunos.all()
    
    # Filtrar atividades por período
    query_atividades = Atividade.query
    
    if turma:
        query_atividades = query_atividades.filter_by(turma=turma)
    
    if data_ini:
        try:
            data_ini_obj = datetime.strptime(data_ini, '%Y-%m-%d')
            query_atividades = query_atividades.filter(Atividade.data_criacao >= data_ini_obj)
        except:
            pass
    
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d')
            query_atividades = query_atividades.filter(Atividade.data_criacao <= data_fim_obj)
        except:
            pass
    
    atividades = query_atividades.all()
    
    # Calcular estatísticas por aluno
    dados_alunos = []
    
    for aluno in alunos:
        entregas = Entrega.query.filter_by(aluno_id=aluno.id).all()
        
        total_entregas = len(entregas)
        entregas_avaliadas = [e for e in entregas if e.nota is not None]
        
        nota_media = sum(float(e.nota) for e in entregas_avaliadas) / len(entregas_avaliadas) if entregas_avaliadas else 0
        
        # Taxa de entrega
        total_atividades = len(atividades)
        taxa_entrega = (total_entregas / total_atividades * 100) if total_atividades > 0 else 0
        
        dados_alunos.append({
            'aluno_id': aluno.id,
            'nome': aluno.nome_completo,
            'email': aluno.email,
            'total_entregas': total_entregas,
            'entregas_avaliadas': len(entregas_avaliadas),
            'nota_media': round(nota_media, 2),
            'taxa_entrega': round(taxa_entrega, 2)
        })
    
    # Ordenar por nota média (ranking)
    dados_alunos_sorted = sorted(dados_alunos, key=lambda x: x['nota_media'], reverse=True)
    
    # Estatísticas gerais
    estatisticas_gerais = {
        'total_alunos': len(alunos),
        'total_atividades': len(atividades),
        'nota_media_turma': round(sum(a['nota_media'] for a in dados_alunos) / len(dados_alunos), 2) if dados_alunos else 0,
        'taxa_entrega_media': round(sum(a['taxa_entrega'] for a in dados_alunos) / len(dados_alunos), 2) if dados_alunos else 0
    }
    
    # Retornar JSON
    if formato == 'json':
        return jsonify({
            'ok': True,
            'estatisticas_gerais': estatisticas_gerais,
            'ranking': dados_alunos_sorted,
            'turma': turma,
            'periodo': {
                'data_inicio': data_ini,
                'data_fim': data_fim
            }
        }), 200
    
    # Gerar XLSX
    elif formato == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Desempenho"
        
        # Cabeçalho
        ws.append(['Relatório de Desempenho - AtivFlow'])
        ws.append([f'Turma: {turma or "Todas"}'])
        ws.append([f'Período: {data_ini or "Início"} a {data_fim or "Fim"}'])
        ws.append([])
        
        # Estatísticas gerais
        ws.append(['Estatísticas Gerais'])
        ws.append(['Total de Alunos', estatisticas_gerais['total_alunos']])
        ws.append(['Total de Atividades', estatisticas_gerais['total_atividades']])
        ws.append(['Nota Média da Turma', estatisticas_gerais['nota_media_turma']])
        ws.append(['Taxa de Entrega Média', f"{estatisticas_gerais['taxa_entrega_media']}%"])
        ws.append([])
        
        # Ranking
        ws.append(['Ranking de Alunos'])
        ws.append(['Posição', 'Nome', 'Email', 'Total Entregas', 'Nota Média', 'Taxa de Entrega (%)'])
        
        for idx, aluno in enumerate(dados_alunos_sorted, 1):
            ws.append([
                idx,
                aluno['nome'],
                aluno['email'],
                aluno['total_entregas'],
                aluno['nota_media'],
                aluno['taxa_entrega']
            ])
        
        # Estilizar cabeçalhos
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14)
        
        for cell in ws[5]:
            cell.font = Font(bold=True)
        
        for cell in ws[11]:
            cell.font = Font(bold=True)
        
        for cell in ws[12]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'relatorio_desempenho_{turma or "todas"}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    # Gerar PDF
    elif formato == 'pdf':
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #003366; }}
                h2 {{ color: #0066cc; margin-top: 30px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #003366; color: white; }}
                .stats {{ background-color: #f0f0f0; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            <h1>Relatório de Desempenho - AtivFlow</h1>
            <p><strong>Turma:</strong> {turma or 'Todas'}</p>
            <p><strong>Período:</strong> {data_ini or 'Início'} a {data_fim or 'Fim'}</p>
            
            <div class="stats">
                <h2>Estatísticas Gerais</h2>
                <p><strong>Total de Alunos:</strong> {estatisticas_gerais['total_alunos']}</p>
                <p><strong>Total de Atividades:</strong> {estatisticas_gerais['total_atividades']}</p>
                <p><strong>Nota Média da Turma:</strong> {estatisticas_gerais['nota_media_turma']}</p>
                <p><strong>Taxa de Entrega Média:</strong> {estatisticas_gerais['taxa_entrega_media']}%</p>
            </div>
            
            <h2>Ranking de Alunos</h2>
            <table>
                <thead>
                    <tr>
                        <th>Posição</th>
                        <th>Nome</th>
                        <th>Email</th>
                        <th>Total Entregas</th>
                        <th>Nota Média</th>
                        <th>Taxa de Entrega (%)</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for idx, aluno in enumerate(dados_alunos_sorted, 1):
            html_content += f"""
                    <tr>
                        <td>{idx}</td>
                        <td>{aluno['nome']}</td>
                        <td>{aluno['email']}</td>
                        <td>{aluno['total_entregas']}</td>
                        <td>{aluno['nota_media']}</td>
                        <td>{aluno['taxa_entrega']}</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        # Gerar PDF
        pdf = HTML(string=html_content).write_pdf()
        
        output = BytesIO(pdf)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'relatorio_desempenho_{turma or "todas"}_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
    
    else:
        return jsonify({'ok': False, 'error': 'Formato inválido'}), 400

